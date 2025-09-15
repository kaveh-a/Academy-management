import flet as ft
import jdatetime
from flet_core.colors import GREEN, WHITE, GREEN_50, GREEN_100, GREEN_200, BLUE_400, BLACK
import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
# from excel_pdf_maker.excel_student_list import create_excel
from db import *
from manager_panel import *
from pathlib import Path

text_color = ft.colors.BLACK

import openpyxl
from datetime import datetime
from tkinter import messagebox
import os



def set_excel_style_simple(worksheet):
    from openpyxl.styles import PatternFill, Border, Side, Alignment
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1:  # هدر
                cell.fill = header_fill
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width

def main(page: ft.Page):
    page.title = "سامانه مدیریت آموزشگاه"
    page.window_width = 1100
    page.window_height = 700
    page.bgcolor = "#E8F5E9"
    page.scroll = ft.ScrollMode.AUTO
    page.window_icon = "app_icon.ico"

    def close_dialog(e):
        page.dialog.open = False
        page.update()

    def show_message(title, message):
        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text(title, color=BLUE_400, weight=ft.FontWeight.BOLD),
            content=ft.Text(message, color=WHITE),
            actions=[
                ft.TextButton("باشه", on_click=close_dialog, style=ft.ButtonStyle(color=GREEN_200))
            ],
        )
        page.dialog = dlg
        dlg.open = True
        page.update()

    def save_and_show(title, message):
        show_message(title, message)

    def show_login():
        page.clean()
        email = ft.TextField(label="نام کاربری", width=300, color=text_color, border_color=ft.colors.GREY_600)
        password = ft.TextField(label="رمز عبور", password=True, width=300, color=text_color, border_color=ft.colors.GREY_600)

        def on_login(e):
            username = email.value.strip()
            passwd = password.value.strip()
            role = check_user_login(username, passwd)
            if not role:
                show_message("خطا", "نام کاربری یا رمز عبور نادرست است")
                return
            if role == "admin":
                show_admin_panel()
            elif role == "manager":
                show_manager_panel(page)
            else:
                show_message("خطا", "دسترسی نامعتبر")

        login_btn = ft.ElevatedButton(
            "ورود",
            width=200,
            height=50,
            bgcolor=ft.colors.GREEN_400,
            color=ft.colors.WHITE,
            on_click=on_login
        )
        page.add(
            ft.Column(
                [
                    ft.Text("ورود به سامانه", size=24, weight=ft.FontWeight.BOLD, color=text_color),
                    ft.Divider(height=20, thickness=1, color=ft.colors.TRANSPARENT),
                    email,
                    password,
                    ft.Divider(height=10),
                    login_btn
                ],
                spacing=20,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                alignment=ft.MainAxisAlignment.CENTER,
                height=page.window_height - 100
            )
        )

    def create_excel():
        try:
            students = get_all_students()

            if not students:
                show_message("اخطار", "هیچ دانش‌آموزی یافت نشد!")
                return

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "لیست دانش‌آموزان"

            headers = ["ردیف", "کد", "نام", "پایه", "موبایل"]
            sheet.append(headers)

            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # آبی حرفه‌ای
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for index, student in enumerate(students, start=1):
                row_data = [
                    index,  # ردیف
                    student.get("student_code", ""),  # کد
                    student.get("student_name", ""),  # نام
                    student.get("student_grade", ""),  # پایه
                    student.get("student_mobile", "")  # موبایل
                ]
                sheet.append(row_data)

            data_alignment = Alignment(horizontal="center", vertical="center")
            data_font = Font(size=11)

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.font = data_font
                    cell.border = thin_border

            column_widths = [8, 12, 25, 15, 18]
            for i, width in enumerate(column_widths, start=1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            filename = f"دانش‌آموزان_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            filepath = os.path.join(desktop_path, filename)
            workbook.save(filepath)

            show_message("اکسل", f"فایل اکسل آماده شد:\n{filename}")

        except Exception as e:
            show_message("خطا", f"خطا در ایجاد فایل اکسل:\n{str(e)}")
            print(f"❌ خطا در create_excel: {e}")

    def set_excel_style(worksheet, start_row_class, start_row_students):

        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        alignment = Alignment(horizontal="center", vertical="center")
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = alignment
                if cell.row == 1:
                    cell.fill = header_fill
                if cell.row == start_row_students and cell.column <= 3:
                    cell.fill = header_fill
                if cell.row > start_row_students and (cell.row - start_row_students) % 2 == 1:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column].width = adjusted_width

    def export_class_to_excel(class_id, class_info, students_list, page):

        try:
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            file_name = f"کلاس_{class_info['class_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(desktop_path, file_name)
            class_data = [
                ["نام کلاس", class_info["class_name"]],
                ["کد کلاس", class_info["class_id"]],
                ["مربی", class_info["teacher_name"]],
                ["زمان", class_info["class_time"]],
                ["تاریخ شروع", class_info["class_start_date"]],
                ["تاریخ پایان", class_info["class_end_date"]],
                ["تعداد دانش‌آموزان", len(students_list)],
            ]
            students_data = [
                [s["student_code"], s["student_name"], s["student_mobile"]] for s in students_list
            ]
            df_class = pd.DataFrame(class_data, columns=["عنوان", "مقدار"])
            df_students = pd.DataFrame(students_data, columns=["کد دانش‌آموز", "نام دانش‌آموز", "موبایل"])
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df_class.to_excel(writer, sheet_name="کلاس", index=False, header=False)
                df_students.to_excel(writer, sheet_name="کلاس", startrow=len(class_data) + 2, index=False)
                worksheet = writer.sheets["کلاس"]
                set_excel_style(worksheet, start_row_class=len(class_data), start_row_students=len(class_data) + 3)
            show_message("موفقیت", f"فایل اکسل روی دسکتاپ ذخیره شد:\n{file_name}")
        except Exception as e:
            show_message("خطا", f"خطا در ایجاد اکسل: {str(e)}")
            print(f"❌ خطا: {e}")

    def show_admin_panel(mode="list"):
        page.clean()

        sidebar = ft.Container(
            content=ft.Column(
                [
                    ft.Text("منوی اصلی", size=18, weight=ft.FontWeight.BOLD, color=WHITE),
                    ft.Divider(),
                    ft.ElevatedButton(
                        "لیست",
                        icon=ft.icons.LIST,
                        width=180,
                        on_click=lambda e: show_admin_panel("list"),
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10), color=WHITE)
                    ),
                    ft.ElevatedButton(
                        "افزودن دانش‌آموز",
                        icon=ft.icons.PERSON_ADD,
                        width=180,
                        on_click=lambda e: show_admin_panel("add"),
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10), color=WHITE)
                    ),
                    ft.ElevatedButton(
                        "ویرایش دانش‌آموز",
                        icon=ft.icons.EDIT,
                        width=180,
                        on_click=lambda e: show_admin_panel("edit"),
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10), color=WHITE)
                    ),
                    ft.ElevatedButton(
                        "جستجو دانش آموز",
                        icon=ft.icons.SEARCH,
                        width=180,
                        on_click=lambda e: show_admin_panel("search"),
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10), color=WHITE)
                    ),
                    ft.ElevatedButton(
                        "ثبت پرداخت",
                        icon=ft.icons.PAYMENT,
                        width=180,
                        on_click=lambda e: show_admin_panel("attendance"),
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10), color=WHITE)
                    ),
                    ft.ElevatedButton(
                        "مدیریت کلاس‌ها",
                        icon=ft.icons.CLASS_,
                        width=180,
                        on_click=lambda e: show_admin_panel("special_classes"),
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10), color=WHITE)
                    ),
                    ft.ElevatedButton(
                        "خروج",
                        icon=ft.icons.LOGOUT,
                        width=180,
                        on_click=lambda e: [save_and_show("خروج", "با موفقیت خارج شدید"), show_login()],
                        bgcolor=ft.colors.RED_400,
                        color=ft.colors.WHITE,
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10))
                    ),
                ],
                spacing=12,
                alignment=ft.MainAxisAlignment.START,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            padding=20,
            bgcolor=ft.colors.GREEN_200,
            border_radius=12,
            width=220,
            height=page.window_height - 80,
            alignment=ft.alignment.top_center,
            border=ft.border.all(1, ft.colors.GREY_300),
            margin=ft.margin.only(right=10)
        )

        content = ft.Container(
            expand=True,
            padding=20,
            bgcolor=ft.colors.GREEN_50,
            border_radius=12,
            alignment=ft.alignment.top_center
        )

        if mode == "list":
            students = get_all_students()
            rows = []
            for idx, s in enumerate(students, start=1):
                rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(str(idx), color=text_color)),
                            ft.DataCell(ft.Text(str(s["student_code"]), color=text_color)),
                            ft.DataCell(ft.Text(s["student_name"], color=text_color)),
                            ft.DataCell(ft.Text(s["student_grade"], color=text_color)),
                            ft.DataCell(ft.Text(s["student_mobile"], color=text_color)),
                        ]
                    )
                )
            table = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("ردیف", color=text_color)),
                    ft.DataColumn(ft.Text("کد", color=text_color)),
                    ft.DataColumn(ft.Text("نام", color=text_color)),
                    ft.DataColumn(ft.Text("پایه", color=text_color)),
                    ft.DataColumn(ft.Text("موبایل", color=text_color)),
                ],
                rows=rows,
                column_spacing=30,
                heading_row_color=ft.colors.GREEN_700,
                data_row_color={"even": "transparent", "odd": ft.colors.GREEN_50},
                border=ft.border.all(1, ft.colors.GREY_400),
                horizontal_lines=ft.BorderSide(1, ft.colors.GREY_300),
                vertical_lines=ft.BorderSide(1, ft.colors.GREY_300),
                width=750,
                height=450,
                show_checkbox_column=False
            )
            export_btn = ft.ElevatedButton(
                "دریافت اکسل",
                icon=ft.icons.DOWNLOAD,
                on_click=lambda e: [create_excel(), show_message("اکسل", "فایل اکسل آماده شد")],
                bgcolor=ft.colors.GREEN_900,
                color=ft.colors.WHITE,
                width=180
            )
            content.content = ft.Column(
                controls=[
                    ft.Text("لیست دانش آموزان", size=22, weight=ft.FontWeight.BOLD, color=text_color),
                    ft.Divider(),
                    export_btn,
                    ft.Divider(),
                    table
                ],
                spacing=20,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER
            )

        elif mode == "add":
            name_field = ft.TextField(label="نام", color=text_color)
            family_field = ft.TextField(label="نام خانوادگی", color=text_color)
            mobile_field = ft.TextField(label="شماره موبایل", color=text_color)
            gender_field = ft.Dropdown(
                label="جنسیت",
                options=[
                    ft.dropdown.Option(text="آقا", key="اقا"),
                    ft.dropdown.Option(text="خانم", key="خانم")
                ]
            )
            grade_field = ft.Dropdown(
                label="پایه تحصیلی",
                options=[
                    ft.dropdown.Option("اول"), ft.dropdown.Option("دوم"), ft.dropdown.Option("سوم"),
                    ft.dropdown.Option("چهارم"), ft.dropdown.Option("پنجم"), ft.dropdown.Option("ششم"),
                    ft.dropdown.Option("هفتم"), ft.dropdown.Option("هشتم"), ft.dropdown.Option("نهم"),
                    ft.dropdown.Option("دهم"), ft.dropdown.Option("یازدهم"), ft.dropdown.Option("دوازدهم")
                ]
            )

            def add_student_handler(e):
                if not name_field.value or not family_field.value:
                    show_message("خطا", "نام و نام خانوادگی الزامی است")
                    return
                success = add_student(
                    name=name_field.value.strip(),
                    family=family_field.value.strip(),
                    mobile=mobile_field.value.strip(),
                    grade=grade_field.value,
                    gender=gender_field.value
                )
                if success:
                    save_and_show("موفقیت", "دانش‌آموز با موفقیت اضافه شد")
                    show_admin_panel("list")  # برگرد به لیست
                else:
                    show_message("خطا", "اضافه کردن دانش‌آموز ناموفق بود")

            form = ft.Column(
                [
                    ft.Text("افزودن دانش‌آموز جدید", size=20, color=text_color),
                    name_field,
                    family_field,
                    mobile_field,
                    gender_field,
                    grade_field,
                    ft.ElevatedButton(
                        "ذخیره",
                        on_click=add_student_handler,
                        bgcolor=ft.colors.GREEN_400,
                        color=ft.colors.WHITE
                    )
                ],
                spacing=15,
                width=400
            )
            content.content = form

        elif mode == "edit":
            search_field = ft.TextField(label="کد دانش‌آموزی", color=text_color)
            name_field = ft.TextField(label="نام", color=text_color)
            family_field = ft.TextField(label="نام خانوادگی", color=text_color)
            mobile_field = ft.TextField(label="شماره موبایل", color=text_color)
            grade_field = ft.Dropdown(
                label="پایه تحصیلی",
                options=[
                    ft.dropdown.Option("اول"), ft.dropdown.Option("دوم"), ft.dropdown.Option("سوم"),
                    ft.dropdown.Option("چهارم"), ft.dropdown.Option("پنجم"), ft.dropdown.Option("ششم"),
                    ft.dropdown.Option("هفتم"), ft.dropdown.Option("هشتم"), ft.dropdown.Option("نهم"),
                    ft.dropdown.Option("دهم"), ft.dropdown.Option("یازدهم"), ft.dropdown.Option("دوازدهم")
                ],
                color=WHITE
            )
            current_code = None

            def find_student(e):
                nonlocal current_code
                try:
                    student = find_student_by_code(int(search_field.value))
                    if student:
                        current_code = student["student_code"]
                        full_name = student["student_name"].split()
                        name_field.value = full_name[0] if len(full_name) > 0 else ""
                        family_field.value = " ".join(full_name[1:]) if len(full_name) > 1 else ""
                        mobile_field.value = student["student_mobile"]
                        grade_field.value = student["student_grade"]
                        page.update()
                        show_message("پیدا شد", "دانش‌آموز با موفقیت پیدا شد")
                    else:
                        show_message("خطا", "دانش‌آموزی با این کد یافت نشد")
                except Exception as ex:
                    show_message("خطا", "کد دانش‌آموز را به درستی وارد کنید")

            def update_student_handler(e):
                if not current_code:
                    return
                if not name_field.value or not family_field.value:
                    show_message("خطا", "نام و نام خانوادگی الزامی است")
                    return
                success = update_student(
                    student_code=current_code,
                    name=name_field.value,
                    family=family_field.value,
                    mobile=mobile_field.value,
                    grade=grade_field.value  # ✅ اضافه شده
                )
                if success:
                    save_and_show("موفقیت", "ویرایش با موفقیت انجام شد")
                    show_admin_panel("list")
                else:
                    show_message("خطا", "ویرایش ناموفق بود")

            form = ft.Column(
                [
                    ft.Text("ویرایش دانش‌آموز", size=20, color=text_color),
                    search_field,
                    ft.ElevatedButton("جستجو", on_click=find_student, bgcolor=ft.colors.BLUE_400,
                                      color=ft.colors.WHITE),
                    ft.Divider(),
                    name_field,
                    family_field,
                    mobile_field,
                    grade_field,  # ✅ اضافه شده
                    ft.ElevatedButton("ذخیره تغییرات", on_click=update_student_handler, bgcolor=ft.colors.GREEN_400,
                                      color=ft.colors.WHITE)
                ],
                spacing=15,
                width=400
            )
            content.content = form

        elif mode == "search":
            search_term = ft.TextField(label="عبارت جستجو", width=250, color=text_color)
            search_type = ft.Dropdown(
                label="جستجو بر اساس",
                options=[ft.dropdown.Option("نام یا نام خانوادگی"), ft.dropdown.Option("کد")],
                width=150,
                color=WHITE
            )
            results_container = ft.Container()
            payments_table = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("ردیف", color=text_color)),
                    ft.DataColumn(ft.Text("مبلغ", color=text_color)),
                    ft.DataColumn(ft.Text("تاریخ", color=text_color)),
                    ft.DataColumn(ft.Text("ساعت", color=text_color)),
                    ft.DataColumn(ft.Text("نوع", color=text_color)),
                    ft.DataColumn(ft.Text("تعداد قسط", color=text_color)),
                    ft.DataColumn(ft.Text("توضیحات", color=text_color)),
                ],
                rows=[],
                column_spacing=20,
                heading_row_color=ft.colors.GREEN_700,
                data_row_color={"even": "transparent", "odd": ft.colors.GREEN_50},
                border=ft.border.all(1, ft.colors.GREY_400),
                width=800,
                height=200,
                show_checkbox_column=False
            )
            payments_section = ft.Container(
                visible=False,
                content=ft.Column([
                    ft.Text("سوابق پرداخت", size=18, weight=ft.FontWeight.BOLD, color=text_color),
                    payments_table
                ]),
                border=ft.border.all(1, ft.colors.GREY_300),
                border_radius=8,
                padding=10,
                margin=ft.margin.only(top=10)
            )
            factor_btn = ft.ElevatedButton(
                "دریافت فاکتور اکسل",
                icon=ft.icons.DOWNLOAD,
                bgcolor=ft.colors.BLUE_400,
                color=ft.colors.WHITE,
                width=220,
                visible=False
            )

            def perform_search(e):
                term = search_term.value.strip()
                by = search_type.value
                if not term or not by:
                    show_message("خطا", "عبارت و نوع جستجو را وارد کنید")
                    return
                students = get_all_students()
                results = []
                if by == "نام یا نام خانوادگی":
                    results = [s for s in students if term in s["student_name"]]
                elif by == "کد":
                    try:
                        code = int(term)
                        results = [s for s in students if s["student_code"] == code]
                    except ValueError:
                        show_message("خطا", "کد دانش‌آموز باید عدد باشد")
                        return

                rows = []
                for idx, s in enumerate(results, start=1):
                    rows.append(
                        ft.DataRow(
                            cells=[
                                ft.DataCell(ft.Text(str(idx), color=text_color)),
                                ft.DataCell(ft.Text(str(s["student_code"]), color=text_color)),
                                ft.DataCell(ft.Text(s["student_name"], color=text_color)),
                                ft.DataCell(ft.Text(s["student_grade"], color=text_color)),
                                ft.DataCell(ft.Text(s["student_mobile"], color=text_color)),
                            ]
                        )
                    )
                table = ft.DataTable(
                    columns=[
                        ft.DataColumn(ft.Text("ردیف", color=text_color)),
                        ft.DataColumn(ft.Text("کد", color=text_color)),
                        ft.DataColumn(ft.Text("نام", color=text_color)),
                        ft.DataColumn(ft.Text("پایه", color=text_color)),
                        ft.DataColumn(ft.Text("موبایل", color=text_color)),
                    ],
                    rows=rows,
                    column_spacing=30,
                    heading_row_color=ft.colors.GREEN_700,
                    data_row_color={"even": "transparent", "odd": ft.colors.GREEN_50},
                    border=ft.border.all(1, ft.colors.GREY_400),
                    horizontal_lines=ft.BorderSide(1, ft.colors.GREY_300),
                    vertical_lines=ft.BorderSide(1, ft.colors.GREY_300),
                    width=750,
                    height=450,
                    show_checkbox_column=False
                )
                results_container.content = ft.Container(content=table, padding=10)

                payments_section.visible = False
                factor_btn.visible = False
                payments_table.rows.clear()

                if len(results) == 1:
                    student_code = results[0]["student_code"]
                    payments = get_payments_by_student_code(student_code)
                    if payments:
                        for idx, p in enumerate(payments, start=1):
                            payments_table.rows.append(
                                ft.DataRow(cells=[
                                    ft.DataCell(ft.Text(str(idx), color=text_color)),
                                    ft.DataCell(ft.Text(str(p["payment_amount"]), color=text_color)),
                                    ft.DataCell(ft.Text(p["payment_date"], color=text_color)),
                                    ft.DataCell(ft.Text(p["payment_time"], color=text_color)),
                                    ft.DataCell(ft.Text(p["payment_type"], color=text_color)),
                                    ft.DataCell(ft.Text(str(p["installments"]) if p["installments"] else "-",
                                                    color=text_color)),
                                    ft.DataCell(ft.Text(p["description"] or "-", color=text_color)),
                                ])
                            )
                        payments_section.visible = True
                        factor_btn.visible = True
                    else:
                        num_cols = len(payments_table.columns)
                        payments_table.rows.append(
                            ft.DataRow(
                                cells=[
                                          ft.DataCell(
                                              ft.Text("پرداختی ثبت نشده", color=ft.colors.GREY_600)
                                          )
                                      ] + [ft.DataCell(ft.Text("")) for _ in range(num_cols - 1)]
                            )
                        )
                        payments_section.visible = True
                        factor_btn.visible = False
                page.update()

            def export_factor_to_excel(e):
                term = search_term.value.strip()
                if not term or search_type.value != "کد":
                    show_message("خطا", "برای فاکتور، کد دانش‌آموز را وارد و جستجو کنید.")
                    return
                try:
                    student_code = int(term)
                    student = None
                    for s in get_all_students():
                        if s["student_code"] == student_code:
                            student = s
                            break
                    if not student:
                        show_message("خطا", "دانش‌آموزی با این کد یافت نشد.")
                        return
                    payments = get_payments_by_student_code(student_code)
                    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
                    file_name = f"فاکتور_{student['student_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    file_path = os.path.join(desktop_path, file_name)
                    student_data = [
                        ["نام", student["student_name"]],
                        ["کد", student["student_code"]],
                        ["پایه", student["student_grade"]],
                        ["موبایل", student["student_mobile"]],
                    ]
                    payment_data = [
                        [p["payment_amount"], p["payment_date"], p["payment_time"], p["payment_type"],
                         p["installments"] or "-", p["description"] or "-"]
                        for p in payments
                    ]
                    df_student = pd.DataFrame(student_data, columns=["عنوان", "مقدار"])
                    df_payments = pd.DataFrame(payment_data,
                                               columns=["مبلغ", "تاریخ", "ساعت", "نوع", "تعداد قسط", "توضیحات"])
                    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                        df_student.to_excel(writer, sheet_name="فاکتور", index=False, header=False)
                        df_payments.to_excel(writer, sheet_name="فاکتور", startrow=6, index=False)
                        worksheet = writer.sheets["فاکتور"]
                        set_excel_style(worksheet, start_row_class=6, start_row_students=7)
                    show_message("موفقیت", f"فاکتور اکسل با موفقیت ذخیره شد:\n{file_name}")
                except Exception as ex:
                    show_message("خطا", f"خطا در ایجاد فاکتور: {str(ex)}")

            factor_btn.on_click = export_factor_to_excel

            content.content = ft.Column([
                ft.Text("جستجوی دانش‌آموز", size=22, weight=ft.FontWeight.BOLD, color=text_color),
                ft.Divider(),
                ft.Row([search_term, search_type,
                        ft.ElevatedButton("جستجو", on_click=perform_search, bgcolor=ft.colors.BLUE_400,
                                          color=ft.colors.WHITE)]),
                ft.Divider(),
                results_container,
                ft.Divider(),
                payments_section,
                factor_btn
            ], spacing=20, horizontal_alignment=ft.CrossAxisAlignment.CENTER)

        elif mode == "attendance":
            student_id = ft.TextField(label="کد دانش‌آموز", width=150, color=text_color)
            amount = ft.TextField(label="مبلغ", width=150, keyboard_type=ft.KeyboardType.NUMBER, color=text_color)
            date = ft.TextField(label="تاریخ", width=150, hint_text="1403/05/01", color=text_color)
            time = ft.TextField(label="ساعت", width=120, hint_text="14:30", color=text_color)
            p_type = ft.Dropdown(
                label="نوع",
                options=[
                    ft.dropdown.Option("کامل"),
                    ft.dropdown.Option("قسطی"),
                    ft.dropdown.Option("شهریه")
                ],
                color=WHITE,
                value="کامل"
            )
            installments_field = ft.TextField(
                label="تعداد قسط",
                width=150,
                keyboard_type=ft.KeyboardType.NUMBER,
                visible=False,
                color=text_color
            )
            description_field = ft.TextField(
                label="توضیحات",
                width=300,
                multiline=True,
                max_lines=2,
                hint_text="مثلاً: مانده 200، پرداخت شده 2 قسط",
                color=text_color
            )

            def on_payment_type_change(e):
                installments_field.visible = (p_type.value == "قسطی")
                page.update()

            p_type.on_change = on_payment_type_change

            def submit_payment(e):
                if not student_id.value or not amount.value or not date.value or not time.value:
                    show_message("خطا", "همه فیلدهای اجباری را پر کنید.")
                    return
                try:
                    student_code = int(student_id.value)
                    amount_val = int(amount.value)
                except ValueError:
                    show_message("خطا", "کد دانش‌آموز و مبلغ باید عدد باشند.")
                    return

                installments_val = installments_field.value if p_type.value == "قسطی" and installments_field.value else None
                description_val = description_field.value.strip() or None

                success = add_payment_to_presence(
                    student_code=student_code,
                    amount=amount_val,
                    date=date.value.strip(),
                    time=time.value.strip(),
                    payment_type=p_type.value,
                    installments=installments_val,
                    description=description_val
                )
                if success:
                    show_message("موفقیت", "پرداخت با موفقیت ثبت شد.")
                    student_id.value = ""
                    amount.value = ""
                    date.value = ""
                    time.value = ""
                    description_field.value = ""
                    if p_type.value != "قسطی":
                        installments_field.value = ""
                    else:
                        installments_field.value = ""
                    p_type.value = "کامل"
                    installments_field.visible = False
                    page.update()
                else:
                    show_message("خطا", "ثبت پرداخت ناموفق بود. دوباره تلاش کنید.")

            content.content = ft.Column([
                ft.Text("ثبت پرداخت", size=22, weight=ft.FontWeight.BOLD, color=BLACK),
                ft.Divider(),
                ft.Row([
                    student_id,
                    amount,
                    date,
                    time,
                    p_type
                ]),
                ft.Row([
                    installments_field,
                    description_field
                ], spacing=10),
                ft.Divider(height=10),
                ft.ElevatedButton(
                    "ثبت فوری پرداخت",
                    on_click=submit_payment,
                    bgcolor=ft.colors.GREEN_400,
                    color=ft.colors.WHITE,
                    width=200,
                    height=50
                )
            ], spacing=15, horizontal_alignment=ft.CrossAxisAlignment.CENTER)

        elif mode == "special_classes":
            search_class_code = ft.TextField(label="کد کلاس", width=180, color=BLACK)
            add_student_code = ft.TextField(label="کد دانش‌آموز", width=180, color=BLACK)
            txt_remove_student = ft.TextField(label="کد دانش‌آموز", width=200, color=BLACK)
            class_name = ft.TextField(label="نام کلاس", color=BLACK, width=250)
            teacher = ft.TextField(label="نام مربی", color=BLACK, width=250)
            time = ft.TextField(label="زمان برگزاری", hint_text="مثلاً دوشنبه 16:00 - 18:00", color=BLACK, width=250)
            start_date = ft.TextField(label="تاریخ شروع", hint_text="1403/06/01", color=BLACK, width=180)
            end_date = ft.TextField(label="تاریخ پایان", hint_text="1403/10/15", color=BLACK, width=180)

            classes_table = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("کد کلاس", color=BLACK)),
                    ft.DataColumn(ft.Text("نام کلاس", color=BLACK)),
                    ft.DataColumn(ft.Text("مربی", color=BLACK)),
                    ft.DataColumn(ft.Text("زمان", color=BLACK)),
                    ft.DataColumn(ft.Text("شروع", color=BLACK)),
                    ft.DataColumn(ft.Text("پایان", color=BLACK)),
                ],
                rows=[],
                column_spacing=20,
                heading_row_color=ft.colors.GREEN_700,
                data_row_color={"even": "transparent", "odd": ft.colors.GREEN_50},
                border=ft.border.all(1, ft.colors.GREY_400),
                width=900,
            )

            enrolled_students_table = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("کد دانش‌آموز", color=BLACK)),
                    ft.DataColumn(ft.Text("نام", color=BLACK)),
                    ft.DataColumn(ft.Text("موبایل", color=BLACK)),
                ],
                rows=[],
                column_spacing=30,
                heading_row_color=ft.colors.BLUE_400,
                data_row_color={"even": "transparent", "odd": ft.colors.BLUE_50},
                border=ft.border.all(1, ft.colors.BLACK),
                width=900,
                height=250,
            )

            current_class_id = None

            def load_classes():
                classes_table.rows.clear()
                classes = get_all_special_classes()
                for c in classes:
                    classes_table.rows.append(
                        ft.DataRow(
                            cells=[
                                ft.DataCell(ft.Text(str(c["class_id"]), color=text_color)),
                                ft.DataCell(ft.Text(c["class_name"], color=text_color)),
                                ft.DataCell(ft.Text(c["teacher_name"], color=text_color)),
                                ft.DataCell(ft.Text(c["class_time"], color=text_color)),
                                ft.DataCell(ft.Text(c["class_start_date"], color=text_color)),
                                ft.DataCell(ft.Text(c["class_end_date"], color=text_color)),
                            ]
                        )
                    )
                page.update()

            def add_class(e):
                if not all([class_name.value, teacher.value, time.value, start_date.value, end_date.value]):
                    show_message("خطا", "همه فیلدها الزامی هستند.")
                    return
                if create_special_class(class_name.value, teacher.value, time.value, start_date.value, end_date.value):
                    show_message("موفق", "کلاس ایجاد شد")
                    class_name.value = ""
                    teacher.value = ""
                    time.value = ""
                    start_date.value = ""
                    end_date.value = ""
                    load_classes()
                else:
                    show_message("خطا", "ایجاد کلاس ناموفق بود")

            def show_enrolled_students(e):
                try:
                    class_id = int(search_class_code.value.strip())
                    nonlocal current_class_id
                    current_class_id = class_id
                    enrolled_students_table.rows.clear()
                    students = get_students_in_class(class_id)
                    if not students:
                        enrolled_students_table.rows.append(
                            ft.DataRow(
                                cells=[
                                    ft.DataCell(
                                        ft.Container(
                                            content=ft.Text(
                                                "هیچ دانش‌آموزی ثبت‌نام نکرده",
                                                color=ft.colors.BLACK,
                                                italic=True,
                                                size=12
                                            ),
                                            alignment=ft.alignment.center,
                                            padding=10,
                                        ),
                                        col_span=3
                                    )
                                ]
                            )
                        )
                    else:
                        for s in students:
                            enrolled_students_table.rows.append(
                                ft.DataRow([
                                    ft.DataCell(ft.Text(str(s["student_code"]), color=ft.colors.BLACK)),
                                    ft.DataCell(ft.Text(s["student_name"], color=ft.colors.BLACK)),
                                    ft.DataCell(ft.Text(s["student_mobile"], color=ft.colors.BLACK)),
                                ])
                            )
                    enrolled_students_table.update()
                    page.update()
                except ValueError:
                    show_message("خطا", "کد کلاس باید عدد باشد")

            def add_student_to_class(e):
                try:
                    class_id = int(search_class_code.value.strip())
                    student_code = int(add_student_code.value.strip())
                except ValueError:
                    show_message("خطا", "کد کلاس و دانش‌آموز باید عدد باشند")
                    return
                if enroll_student_in_class(class_id, student_code):
                    show_message("موفق", "دانش‌آموز به کلاس اضافه شد")
                    add_student_code.value = ""
                    show_enrolled_students(None)
                else:
                    show_message("خطا", "ثبت نام ناموفق بود (قبلاً ثبت‌نام کرده است)")

            def remove_student_from_class(e):
                conn = None
                try:
                    student_code_str = txt_remove_student.value.strip()
                    if not student_code_str:
                        show_message("❌ خطا", "لطفاً کد دانش‌آموز را وارد کنید.")
                        return
                    if not student_code_str.isdigit():
                        show_message("❌ خطا", "کد دانش‌آموز باید فقط عدد انگلیسی باشد (مثلاً 123).")
                        return
                    student_code = int(student_code_str)
                    if not current_class_id:
                        show_message("❌ خطا", "لطفاً ابتدا کد کلاس را مشاهده کنید.")
                        return
                    conn = sqlite3.connect("nihan_danesh.db")
                    cursor = conn.cursor()
                    cursor.execute("""
                        DELETE FROM class_enrollments 
                        WHERE class_id = ? AND student_code = ?
                    """, (current_class_id, student_code))
                    if cursor.rowcount == 0:
                        show_message("⚠️ هشدار", f"دانش‌آموز با کد {student_code} در این کلاس وجود ندارد.")
                    else:
                        conn.commit()
                        show_message("✅ موفق", f"دانش‌آموز با کد {student_code} حذف شد.")
                        show_enrolled_students(None)
                        txt_remove_student.value = ""
                except Exception as ex:
                    show_message("❌ خطا", f"خطای غیرمنتظره: {str(ex)}")
                finally:
                    if conn:
                        conn.close()
                page.update()

            def export_class_excel(e):
                try:
                    selected_class_id = int(search_class_code.value.strip())
                except (ValueError, TypeError):
                    show_message("❌ خطا", "لطفاً یک کد کلاس معتبر وارد کنید.")
                    return

                desktop_path = Path.home() / "Desktop"
                output_file = desktop_path / f"اطلاعات_کلاس_{selected_class_id}.xlsx"

                conn = None
                try:
                    conn = sqlite3.connect("nihan_danesh.db")
                    cursor = conn.cursor()

                    cursor.execute("""
                        SELECT class_id, class_name, teacher_name, class_time, class_start_date, class_end_date
                        FROM special_classes WHERE class_id = ?
                    """, (selected_class_id,))
                    class_row = cursor.fetchone()
                    if not class_row:
                        show_message("❌ خطا", "کلاسی با این کد یافت نشد.")
                        return

                    columns = ['class_id', 'class_name', 'teacher_name', 'class_time', 'class_start_date', 'class_end_date']
                    class_info = dict(zip(columns, class_row))

                    cursor.execute("""
                        SELECT s.student_code, s.student_name, s.student_family, s.student_mobile, s.student_grade
                        FROM students s
                        INNER JOIN class_enrollments ce ON s.student_code = ce.student_code
                        WHERE ce.class_id = ?
                    """, (selected_class_id,))
                    student_rows = cursor.fetchall()
                    student_columns = ['کد دانش‌آموزی', 'نام', 'نام خانوادگی', 'شماره موبایل', 'پایه']
                    df_students = pd.DataFrame(student_rows, columns=student_columns)
                    df_students['نام کامل'] = df_students['نام'] + " " + df_students['نام خانوادگی']
                    df_students = df_students[['کد دانش‌آموزی', 'نام کامل', 'شماره موبایل', 'پایه']]

                    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                        start_row = 0
                        class_data = [
                            ["نام کلاس", class_info['class_name']],
                            ["مربی", class_info['teacher_name']],
                            ["زمان کلاس", class_info['class_time']],
                            ["تاریخ شروع", class_info['class_start_date']],
                            ["تاریخ پایان", class_info['class_end_date']],
                            ["کد کلاس", class_info['class_id']]
                        ]
                        df_class = pd.DataFrame(class_data, columns=["ویژگی", "مقدار"])
                        df_class.to_excel(writer, sheet_name='کلاس', startrow=start_row, index=False)

                        worksheet = writer.sheets['کلاس']
                        workbook = writer.book

                        center_format = workbook.add_format({
                            'align': 'center',
                            'valign': 'vcenter',
                            'font_name': 'B Nazanin',
                            'font_size': 11
                        })
                        header_format = workbook.add_format({
                            'bold': True,
                            'bg_color': '#D3D3D3',
                            'align': 'center',
                            'valign': 'vcenter',
                            'font_name': 'B Titr',
                            'font_size': 12,
                            'border': 1
                        })

                        for col_num in range(len(df_class.columns)):
                            worksheet.set_column(col_num, col_num, 20, center_format)
                        for col_num, value in enumerate(df_class.columns):
                            worksheet.write(start_row, col_num, value, header_format)

                        start_row += len(df_class) + 3

                        if not df_students.empty:
                            df_students.to_excel(writer, sheet_name='کلاس', startrow=start_row, index=False)
                            for col_num, value in enumerate(df_students.columns):
                                worksheet.write(start_row, col_num, value, header_format)
                            for col_num in range(len(df_students.columns)):
                                max_len = max(
                                    df_students.iloc[:, col_num].astype(str).map(len).max(),
                                    len(df_students.columns[col_num])
                                ) + 2
                                worksheet.set_column(col_num, col_num, max_len, center_format)
                        else:
                            worksheet.write(start_row, 0, "هیچ دانش‌آموزی در این کلاس ثبت نشده است.", center_format)
                            worksheet.set_column(0, 0, 40)

                        worksheet.set_default_row(20)

                    show_message("✅ موفق", f"فایل اکسل با موفقیت ذخیره شد:\n{output_file}")
                except Exception as ex:
                    show_message("❌ خطا", f"خطا در ایجاد اکسل: {str(ex)}")
                finally:
                    if conn:
                        conn.close()

            content.content = ft.Column(
                [
                    ft.Text("مدیریت کلاس‌های تخصصی", size=22, weight=ft.FontWeight.BOLD, color=text_color),
                    ft.Divider(),
                    # ایجاد کلاس جدید
                    ft.Text("ایجاد کلاس جدید", size=18, weight=ft.FontWeight.BOLD, color=text_color),
                    ft.Row([class_name, teacher], spacing=10),
                    ft.Row([time, start_date, end_date], spacing=10),
                    ft.ElevatedButton("افزودن کلاس", on_click=add_class, bgcolor=ft.colors.GREEN_400,
                                      color=ft.colors.WHITE),
                    ft.Divider(height=20),
                    ft.Container(
                        content=classes_table,
                        border=ft.border.all(1, ft.colors.GREY_300),
                        border_radius=8,
                        padding=10
                    ),
                    ft.Divider(height=20),
                    ft.Text("مشاهده دانش‌آموزان کلاس", size=18, weight=ft.FontWeight.BOLD, color=text_color),
                    ft.Row([
                        search_class_code,
                        ft.ElevatedButton("مشاهده", on_click=show_enrolled_students, bgcolor=ft.colors.BLUE_400,
                                          color=ft.colors.WHITE),
                    ], spacing=10),
                    ft.Container(
                        content=enrolled_students_table,
                        border=ft.border.all(1, ft.colors.GREY_300),
                        border_radius=8,
                        padding=10
                    ),
                    ft.Divider(height=20),
                    ft.ElevatedButton(
                        "دریافت اکسل",
                        icon=ft.icons.DOWNLOAD,
                        on_click=export_class_excel,
                        bgcolor=ft.colors.GREEN_900,
                        color=ft.colors.WHITE
                    ),
                    ft.Divider(height=20),
                    ft.Text("افزودن دانش‌آموز به کلاس", size=18, weight=ft.FontWeight.BOLD, color=text_color),
                    ft.Row([
                        add_student_code,
                        ft.ElevatedButton("افزودن", on_click=add_student_to_class, bgcolor=ft.colors.GREEN_400,
                                          color=ft.colors.WHITE),
                    ], spacing=10),
                    ft.Divider(height=20),
                    ft.Text("حذف دانش‌آموز از کلاس", size=18, weight=ft.FontWeight.BOLD, color=text_color),
                    ft.Text("کد کلاس را در قسمت 'مشاهده دانش‌آموزان' وارد کرده‌اید. حالا کد دانش‌آموز را وارد کنید:",
                            size=12, color=ft.colors.GREY),
                    ft.Row([
                        txt_remove_student,
                        ft.ElevatedButton(
                            "حذف دانش‌آموز",
                            on_click=remove_student_from_class,
                            bgcolor=ft.colors.RED_400,
                            color=ft.colors.WHITE
                        ),
                    ], spacing=10),
                ],
                spacing=20,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                scroll=ft.ScrollMode.AUTO
            )

            load_classes()

        page.add(
            ft.Row(
                [
                    content,
                    sidebar
                ],
                spacing=10,
                alignment=ft.MainAxisAlignment.START,
                vertical_alignment=ft.CrossAxisAlignment.START,
                expand=True
            )
        )
        page.update()

    show_login()

if __name__ == "__main__":
    init_db()
    ft.app(target=main)